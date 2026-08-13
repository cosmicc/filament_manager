"""One-time workbook analysis and explicitly approved PostgreSQL import."""

import hashlib
import json
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from filament_manager.domain.colors import normalize_color_name
from filament_manager.domain.profile_inheritance import (
    profile_columns_from_settings,
    sparse_profile_overrides,
)
from filament_manager.models.auth import User
from filament_manager.models.enums import ProfileStatus, SpoolStatus, UserRole
from filament_manager.models.inventory import (
    FilamentColor,
    FilamentProduct,
    MaterialProfile,
    MaterialTemplate,
    MaterialTemplateRevision,
    Printer,
    Spool,
    Vendor,
)
from filament_manager.models.operations import ImportRun

from .events import add_audit_event, add_outbox_job

HEADERS = [
    "Spool ID",
    "Inventory Status",
    "Material Type",
    "Filler / Reinforcement",
    "Finish / Effect",
    "Color",
    "Manufacturer",
    "Product / Grade / Hardness",
    "Preferred Brand",
    "Diameter (mm)",
    "Tolerance (mm)",
    "Density (g/cm³)",
    "Nominal Weight (g)",
    "Empty Spool / Tare (g)",
    "Current Gross Weight (g)",
    "Remaining Filament (g)",
    "Remaining %",
    "Used Filament (g)",
    "Est. Remaining Length (m)",
    "Full Spool Weight (g)",
    "Flow (%)",
    "Nozzle Temp (°C)",
    "Bed Temp (°C)",
    "Chamber Temp (°C)",
    "Retraction Distance (mm)",
    "Retraction Speed (mm/s)",
    "Pressure Advance",
    "Purchase Source",
    "Purchase Date",
    "First Used",
    "Last Used",
    "Spool Cost",
    "Cost / Gram",
    "Notes",
]


def file_sha256(path: Path) -> str:
    """Hash the exact workbook bytes used for approval."""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _decimal(value: Any, field: str, errors: list[str], *, positive: bool = False) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        errors.append(f"{field} must be numeric")
        return None
    if positive and number <= 0:
        errors.append(f"{field} must be positive")
    if not positive and number < 0:
        errors.append(f"{field} must be non-negative")
    return number


@dataclass(frozen=True)
class AnalyzedRow:
    """Normalized workbook row plus its row-specific validation findings."""

    row_number: int
    values: dict[str, Any]
    errors: list[str]
    warnings: list[str]


def analyze_workbook(path: Path, *, source_name: str | None = None) -> dict[str, Any]:
    """Validate every populated inventory row without changing external state."""

    workbook = load_workbook(path, data_only=True, read_only=True)
    if "Inventory" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("workbook does not contain the Inventory sheet")
    sheet = workbook["Inventory"]
    actual_headers = [sheet.cell(1, column).value for column in range(1, len(HEADERS) + 1)]
    if actual_headers != HEADERS:
        workbook.close()
        raise ValueError("Inventory headers do not match the supplied 34-column contract")

    rows: list[AnalyzedRow] = []
    seen: set[str] = set()
    row_iterator = sheet.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True)
    for row_number, row_values in enumerate(row_iterator, start=2):
        code = row_values[0]
        if code in (None, ""):
            continue
        values = dict(zip(HEADERS, row_values, strict=True))
        errors: list[str] = []
        warnings: list[str] = []
        normalized_code = str(code).strip().upper()
        if normalized_code in seen:
            errors.append("Spool ID is duplicated")
        seen.add(normalized_code)
        values["Spool ID"] = normalized_code
        if not values["Material Type"]:
            errors.append("Material Type is required")
        if not values["Color"]:
            errors.append("Color is required")
        diameter = _decimal(values["Diameter (mm)"], "Diameter", errors, positive=True)
        density = _decimal(values["Density (g/cm³)"], "Density", errors, positive=True)
        nominal = _decimal(values["Nominal Weight (g)"], "Nominal weight", errors, positive=True)
        tare = _decimal(values["Empty Spool / Tare (g)"], "Tare", errors)
        gross = _decimal(values["Current Gross Weight (g)"], "Gross weight", errors)
        cost = _decimal(values["Spool Cost"], "Spool cost", errors)
        if gross is not None and tare is None:
            errors.append("Tare is required when gross weight is present")
        if gross is not None and tare is not None and gross < tare:
            errors.append("Gross weight cannot be below tare")
        if gross is not None and tare is not None and nominal is not None and gross - tare > nominal:
            warnings.append("Measured net mass exceeds nominal capacity and needs administrator review")
        if tare is None:
            warnings.append("Tare is unknown; weighing is blocked until tare is entered")
        values.update(
            {
                "_diameter": diameter,
                "_density": density,
                "_nominal": nominal,
                "_tare": tare,
                "_gross": gross,
                "_cost": cost,
            }
        )
        rows.append(AnalyzedRow(row_number, values, errors, warnings))

    workbook.close()
    return {
        "source": source_name or path.name,
        "sha256": file_sha256(path),
        "inventory_columns": len(HEADERS),
        "populated_rows": len(rows),
        "valid_rows": sum(not row.errors for row in rows),
        "invalid_rows": sum(bool(row.errors) for row in rows),
        "rows": [
            {
                "row_number": row.row_number,
                "spool_code": row.values["Spool ID"],
                "errors": row.errors,
                "warnings": row.warnings,
            }
            for row in rows
        ],
    }


async def save_dry_run(
    session: AsyncSession,
    path: Path,
    *,
    source_name: str | None = None,
    run_id: UUID | None = None,
) -> ImportRun:
    """Persist a hash-bound dry-run report for explicit later approval."""

    report = analyze_workbook(path, source_name=source_name)
    run = ImportRun(
        id=run_id or uuid4(),
        source_name=source_name or path.name,
        source_sha256=str(report["sha256"]),
        dry_run=True,
        status="validated" if report["invalid_rows"] == 0 else "invalid",
        report=report,
        created_at=datetime.now(UTC),
        completed_at=datetime.now(UTC),
    )
    session.add(run)
    await session.commit()
    return run


async def commit_approved_run(
    session: AsyncSession,
    *,
    run_id: UUID,
    workbook_path: Path,
    administrator_username: str,
    audit_source: str = "cli",
    correlation_id: str | None = None,
) -> dict[str, int]:
    """Import the exact approved workbook hash in one canonical transaction."""

    run = await session.scalar(select(ImportRun).where(ImportRun.id == run_id).with_for_update())
    if run is None or not run.dry_run:
        raise ValueError("approved dry-run report was not found")
    if run.status != "validated":
        raise ValueError("only a validated dry run can be approved")
    if run.source_sha256 != file_sha256(workbook_path):
        raise ValueError("workbook changed after dry-run approval")
    administrator = await session.scalar(
        select(User).where(func.lower(User.normalized_username) == administrator_username.casefold())
    )
    if administrator is None or administrator.role != UserRole.ADMINISTRATOR:
        raise ValueError("an active administrator must approve the import")
    if await session.scalar(select(func.count(Spool.id))):
        raise ValueError("canonical spool inventory is not empty")
    printer = await session.scalar(select(Printer).order_by(Printer.created_at).limit(1))
    if printer is None:
        raise ValueError("seed the configured printer before importing profiles")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook["Inventory"]
    vendors: dict[str, Vendor] = {}
    products: dict[tuple[object, ...], FilamentProduct] = {}
    colors: dict[str, FilamentColor] = {}
    imported_spools = 0
    imported_profiles = 0
    template_revisions: dict[tuple[str, UUID, Decimal], MaterialTemplateRevision] = {}
    row_iterator = sheet.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True)
    for _row_number, row_values in enumerate(row_iterator, start=2):
        values = dict(zip(HEADERS, row_values, strict=True))
        if not values["Spool ID"]:
            continue
        vendor_name = str(values["Manufacturer"] or "Unknown").strip()
        vendor = vendors.get(vendor_name.casefold())
        if vendor is None:
            vendor = Vendor(
                name=vendor_name,
                preferred=str(values["Preferred Brand"] or "").casefold() == "yes",
            )
            session.add(vendor)
            await session.flush()
            vendors[vendor_name.casefold()] = vendor
        product_key = (
            vendor.id,
            values["Material Type"],
            values["Filler / Reinforcement"],
            values["Finish / Effect"],
            values["Color"],
            values["Product / Grade / Hardness"],
            str(values["Diameter (mm)"]),
            str(values["Density (g/cm³)"]),
            str(values["Nominal Weight (g)"]),
        )
        product = products.get(product_key)
        if product is None:
            color_name = str(values["Color"]).strip()
            normalized_color = normalize_color_name(color_name)
            color = colors.get(normalized_color)
            if color is None:
                color = await session.scalar(
                    select(FilamentColor).where(FilamentColor.normalized_name == normalized_color)
                )
            if color is None:
                color = FilamentColor(
                    name=color_name,
                    normalized_name=normalized_color,
                    color_hex="808080",
                )
                session.add(color)
                await session.flush()
            colors[normalized_color] = color
            product = FilamentProduct(
                vendor_id=vendor.id,
                material_type=str(values["Material Type"]),
                filler=str(values["Filler / Reinforcement"] or "None"),
                finish=str(values["Finish / Effect"] or "Standard"),
                color_name=color.name,
                color_hex=color.color_hex,
                product_name=str(values["Product / Grade / Hardness"] or "") or None,
                diameter_mm=Decimal(str(values["Diameter (mm)"])),
                tolerance_mm=(
                    Decimal(str(values["Tolerance (mm)"])) if values["Tolerance (mm)"] is not None else None
                ),
                density_g_cm3=Decimal(str(values["Density (g/cm³)"])),
                nominal_net_mass_g=Decimal(str(values["Nominal Weight (g)"])),
            )
            session.add(product)
            await session.flush()
            products[product_key] = product
        nominal = Decimal(str(values["Nominal Weight (g)"]))
        tare = Decimal(str(values["Empty Spool / Tare (g)"] or 0))
        gross = values["Current Gross Weight (g)"]
        measured = Decimal(str(gross)) - tare if gross is not None else None
        spool = Spool(
            spool_code=str(values["Spool ID"]).strip().upper(),
            filament_product_id=product.id,
            nominal_net_mass_g=nominal,
            tare_mass_g=tare,
            remaining_mass_expected_g=measured if measured is not None else nominal,
            remaining_mass_measured_g=measured,
            remaining_mass_effective_g=measured if measured is not None else nominal,
            weight_confidence="measured" if measured is not None else "unknown_tare",
            status=SpoolStatus.IN_STOCK if measured is not None else SpoolStatus.NEEDS_WEIGHING,
            purchase_source=values["Purchase Source"],
            purchase_date=(
                values["Purchase Date"].date()
                if isinstance(values["Purchase Date"], datetime)
                else values["Purchase Date"]
            ),
            purchase_cost=Decimal(str(values["Spool Cost"])) if values["Spool Cost"] is not None else None,
            first_used_at=values["First Used"],
            last_used_at=values["Last Used"],
            notes=values["Notes"],
        )
        session.add(spool)
        await session.flush()
        imported_spools += 1
        add_outbox_job(
            session,
            job_type="spoolman.spool.upsert",
            idempotency_key=f"import:spool:{spool.id}:v1",
            aggregate_type="spool",
            aggregate_id=spool.id,
            aggregate_version=1,
            payload={"spool_id": str(spool.id)},
        )
        if product.id and not await session.scalar(
            select(MaterialProfile.id).where(
                MaterialProfile.filament_product_id == product.id,
                MaterialProfile.printer_id == printer.id,
            )
        ):
            if values["Nozzle Temp (°C)"] is not None and values["Bed Temp (°C)"] is not None:
                profile_settings: dict[str, object] = {
                    "chamber_temp_c": str(values["Chamber Temp (°C)"] or 0),
                    "extruder_temp_c": str(values["Nozzle Temp (°C)"]),
                    "bed_temp_c": str(values["Bed Temp (°C)"]),
                    "flow_percent": str(values["Flow (%)"] or 100),
                    "print_speed_mm_s": None,
                    "outer_wall_speed_mm_s": None,
                    "inner_wall_speed_mm_s": None,
                    "infill_speed_mm_s": None,
                    "top_bottom_speed_mm_s": None,
                    "initial_layer_speed_mm_s": None,
                    "travel_speed_mm_s": None,
                    "support_speed_mm_s": None,
                    "retraction_distance_mm": str(values["Retraction Distance (mm)"] or 0),
                    "retraction_speed_mm_s": str(values["Retraction Speed (mm/s)"] or 0),
                    "cooling_enabled": True,
                    "cooling_min_percent": "0",
                    "cooling_max_percent": "100",
                    "support_overhang_angle_deg": None,
                    "tree_max_branch_angle_deg": None,
                    "pressure_advance": (
                        str(values["Pressure Advance"]) if values["Pressure Advance"] is not None else None
                    ),
                    "filament_density_g_cm3": str(product.density_g_cm3),
                    "preferred_build_plate_surface_id": None,
                    "cura_extensions": {},
                }
                template_key = (
                    product.material_type.casefold(),
                    printer.id,
                    printer.nozzle_diameter_mm,
                )
                template_revision = template_revisions.get(template_key)
                if template_revision is None:
                    template_revision = await session.scalar(
                        select(MaterialTemplateRevision)
                        .join(
                            MaterialTemplate,
                            MaterialTemplate.id == MaterialTemplateRevision.material_template_id,
                        )
                        .where(
                            func.lower(MaterialTemplate.material_type) == product.material_type.casefold(),
                            MaterialTemplate.printer_id == printer.id,
                            MaterialTemplate.nozzle_diameter_mm == printer.nozzle_diameter_mm,
                            MaterialTemplateRevision.status == ProfileStatus.PUBLISHED,
                        )
                        .order_by(MaterialTemplateRevision.version.desc())
                        .limit(1)
                    )
                if template_revision is None:
                    template = MaterialTemplate(
                        name=f"Template {product.material_type}",
                        material_type=product.material_type,
                        description="Created from the approved initial inventory import.",
                        printer_id=printer.id,
                        nozzle_diameter_mm=printer.nozzle_diameter_mm,
                        filament_diameter_mm=product.diameter_mm,
                        active=True,
                    )
                    session.add(template)
                    await session.flush()
                    checksum_payload = {
                        "material_template_id": str(template.id),
                        "version": 1,
                        "settings": profile_settings,
                    }
                    template_revision = MaterialTemplateRevision(
                        material_template_id=template.id,
                        version=1,
                        status=ProfileStatus.PUBLISHED,
                        settings=profile_settings,
                        checksum=hashlib.sha256(
                            json.dumps(
                                checksum_payload,
                                sort_keys=True,
                                separators=(",", ":"),
                            ).encode("utf-8")
                        ).hexdigest(),
                        published_at=datetime.now(UTC),
                    )
                    session.add(template_revision)
                    await session.flush()
                template_revisions[template_key] = template_revision
                product.source_template_revision_id = template_revision.id
                profile = MaterialProfile(
                    **profile_columns_from_settings(profile_settings),
                    filament_product_id=product.id,
                    printer_id=printer.id,
                    nozzle_diameter_mm=printer.nozzle_diameter_mm,
                    version=1,
                    status=ProfileStatus.DRAFT,
                    base_template_revision_id=template_revision.id,
                    setting_overrides=sparse_profile_overrides(
                        template_revision.settings,
                        profile_settings,
                    ),
                )
                session.add(profile)
                imported_profiles += 1

    workbook.close()
    run.status = "committed"
    run.approved_by = administrator.id
    run.completed_at = datetime.now(UTC)
    run.report = {**run.report, "committed_spools": imported_spools, "committed_profiles": imported_profiles}
    add_outbox_job(
        session,
        job_type="google.inventory.publish",
        idempotency_key=f"import:{run.id}:google:v1",
        aggregate_type="import_run",
        aggregate_id=run.id,
        aggregate_version=1,
        payload={"import_run_id": str(run.id)},
    )
    add_audit_event(
        session,
        actor_id=administrator.id,
        source=audit_source,
        action="workbook.import.commit",
        object_type="import_run",
        object_id=run.id,
        before={"status": "validated"},
        after={"status": "committed", "spools": imported_spools, "profiles": imported_profiles},
        correlation_id=correlation_id or f"import-{run.id}",
    )
    await session.commit()
    return {
        "spools": imported_spools,
        "profiles": imported_profiles,
        "vendors": len(vendors),
        "products": len(products),
    }
